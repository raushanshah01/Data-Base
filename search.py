from openpyxl import load_workbook
import pandas as pd

def search():
    workbook = load_workbook(filename="studentdata.xlsx")
    sheet = workbook.active

    search_id = int(input("Enter student ID: "))

    found = False

    for row in sheet.iter_rows(min_row=2, 
                            min_col=1,  
                            max_col=9,  
                            values_only=True):
        student_id = row[0]
        if student_id == search_id:
            data = {
                "reg_no": row[0],
                "name": row[1],
                "phone_no":row[2],
                "Math": row[3],
                "Physics": row[4],
                "Chemistry": row[5],
                "Hindi": row[6],
                "English": row[7],
                "cgpa": row[8]
            }
            found = True
            with open("results.txt", "w") as f:
                df=pd.DataFrame([data])
                f.write(df.to_string(index=False))
            print("Student found.")
            print(df.to_string(index=False))
        
            break  

    if not found:
        print("Student not found.")