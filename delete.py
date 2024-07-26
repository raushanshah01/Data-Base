from openpyxl import load_workbook

def delete():
    workbook = load_workbook(filename="studentdata.xlsx")
    sheet = workbook.active

    search_id = int(input("Enter student ID: "))

    found = False

    for row in range(2, sheet.max_row + 1):
        student_id = sheet.cell(row, 1).value
        if student_id == search_id:
            found = True
            sheet.delete_rows(row,1)
            workbook.save("studentdata.xlsx")
            print("Student details deleted.")
            break

    if not found:
        print("Student not found.")

