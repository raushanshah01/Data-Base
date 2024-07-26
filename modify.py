from openpyxl import load_workbook
import sys

def modify_data():
    workbook = load_workbook(filename="studentdata.xlsx")
    sheet = workbook.active

    search_id = int(input("Enter student ID: "))

    found_row = None

    for row in range(2, sheet.max_row + 1):
        student_id = sheet.cell(row=row, column=1).value
        if student_id == search_id:
            found_row = row
            found = True  # This line is unnecessary if we use `found_row`

            # Initialize variables here
            name = sheet.cell(row=row, column=2).value
            phone_no = sheet.cell(row=row, column=3).value
            marks = {
                "Math": sheet.cell(row=row, column=4).value,
                "Physics": sheet.cell(row=row, column=5).value,
                "Chemistry": sheet.cell(row=row, column=6).value,
                "Hindi": sheet.cell(row=row, column=7).value,
                "English": sheet.cell(row=row, column=8).value
            }

            while True:
                sys.stdout.write("\nModify Student data:\n")
                sys.stdout.write("1. Reg. no.\n")
                sys.stdout.write("2. Name\n")
                sys.stdout.write("3. Phone no.\n")
                sys.stdout.write("4. Marks\n")
                sys.stdout.write("5. Exit...\n")
                choice = input("Enter your choice: ")
                if choice == '1':
                    student_id = int(input("Enter Reg. no:"))
                elif choice == '2':
                    name = input("Enter Student name:")
                elif choice == '3':
                    phone_no = input("Enter Student phone no.:")  # Change this to handle string
                elif choice == '4':
                    for subject in marks.keys():
                        while True:
                            try:
                                mark = int(input(f"Enter {subject} marks: "))
                                if mark < 0 or mark > 100:
                                    raise ValueError
                                marks[subject] = mark
                                break
                            except ValueError:
                                print("Invalid marks. Please enter a value between 0 and 100.")
                elif choice == '5':
                    sys.stdout.write("Exiting...\n")
                    break
                else:
                    sys.stdout.write("Invalid choice. Please choose again.\n")

            total_marks = sum(marks.values())
            average_marks = total_marks / len(marks)
            grade = average_marks / 10  # Grade calculation might be specific to your system

            # Update the values in the current row
            sheet.cell(row=found_row, column=1).value = student_id
            sheet.cell(row=found_row, column=2).value = name
            sheet.cell(row=found_row, column=3).value = phone_no
            col = 4
            for subject, mark in marks.items():
                sheet.cell(row=found_row, column=col).value = mark
                col += 1
            sheet.cell(row=found_row, column=col).value = grade

            workbook.save("studentdata.xlsx")
            print("Modification completed.")
            break

    if not found_row:
        print("Student not found.")

