from openpyxl import Workbook, load_workbook

def add_student():
    try:
        wb = load_workbook(filename="studentdata.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        wb.save("studentdata.xlsx")

    sheet = wb.active

    if sheet.max_row == 1:
        header_row = ["Registration No.", "Name", "Phone Number", "Math", "Physics", "Chemistry", "Hindi", "English", "CGPA"]
        sheet.append(header_row)

    number_of_students = int(input("Number of students to add: "))

    for _ in range(number_of_students):
        student_id = int(input("Enter student ID: "))
        name = input("Enter student name: ")
        phone_no = input("Enter phone number: ")

        marks = {}
        for subject in ["Math", "Physics", "Chemistry", "Hindi", "English"]:
            while True:
                try:
                    marks[subject] = int(input(f"Enter {subject} marks: "))
                    if marks[subject] < 0 or marks[subject] > 100:
                        raise ValueError
                    break
                except ValueError:
                    print("Invalid marks. Please enter a value between 0 and 100.")

        total_marks = sum(marks.values())
        average_marks = total_marks / len(marks)
        grade = average_marks / 10

        student_data = [student_id, name, phone_no, *marks.values(), grade]
        sheet.append(student_data)

    wb.save("studentdata.xlsx")
    print("Excel file with student details updated successfully.")

add_student()